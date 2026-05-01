"""
Excel ↔ country-profile JSON conversion utilities.

Sheet layout (PyPSA-inspired):
  country          – two-column key/value table
  generators       – one row per generator, basic economic params as columns
  cf_eff_func      – one row per generator, function type + flattened params
  eta_func         – same
  integration_cost_func – same
  ess              – two-column key/value table for storage params
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

GENERATOR_BASIC_COLS: list[str] = [
    "capex_usd_kw",
    "opex_fixed_usd_kw_yr",
    "opex_var_usd_mwh",
    "lifetime_yr",
    "emission_factor_tco2_mwh",
    "fuel_usd_mmbtu",
    "heat_rate_mmbtu_mwh",
    "cf_base",
    "variability_factor",
]

# All possible function param keys across all function types
FUNC_PARAM_COLS: list[str] = ["a", "b", "c", "intercept", "threshold", "slope_before", "slope_after"]

FUNC_SHEETS: list[str] = ["cf_eff_func", "eta_func", "integration_cost_func", "curtailment_func"]

# Styling
HEADER_FILL = PatternFill("solid", fgColor="1E293B")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="E2E8F0")
SUBHEADER_FONT = Font(bold=True)


def _style_header(cell: Any) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center")


def _style_subheader(cell: Any) -> None:
    cell.fill = SUBHEADER_FILL
    cell.font = SUBHEADER_FONT


def _autofit(ws: Any, padding: int = 2) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + padding


# ---------------------------------------------------------------------------
# JSON → Workbook
# ---------------------------------------------------------------------------

def profile_to_workbook(profile: dict) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    _write_country_sheet(wb, profile)
    _write_generators_sheet(wb, profile)
    for func_key in FUNC_SHEETS:
        _write_func_sheet(wb, profile, func_key)
    _write_ess_sheet(wb, profile)

    return wb


def _write_country_sheet(wb: Workbook, profile: dict) -> None:
    ws = wb.create_sheet("country")
    headers = ["parameter", "value"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        _style_header(cell)

    rows = [
        ("name", profile.get("name", "")),
        ("annual_generation_twh", profile.get("annual_generation_twh", "")),
        ("discount_rate", profile.get("discount_rate", "")),
    ]
    sources = profile.get("sources", [])
    for i, src in enumerate(sources):
        rows.append((f"source_{i+1}", src))

    for r, (k, v) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)

    _autofit(ws)


def _write_generators_sheet(wb: Workbook, profile: dict) -> None:
    ws = wb.create_sheet("generators")
    all_cols = ["generator"] + GENERATOR_BASIC_COLS

    # Header row
    for col, h in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col, value=h)
        _style_header(cell)

    generators: dict = profile.get("generators", {})
    for r, (gen_name, gen_cfg) in enumerate(generators.items(), 2):
        ws.cell(row=r, column=1, value=gen_name)
        for col, field in enumerate(GENERATOR_BASIC_COLS, 2):
            ws.cell(row=r, column=col, value=gen_cfg.get(field, ""))

    _autofit(ws)


def _write_func_sheet(wb: Workbook, profile: dict, func_key: str) -> None:
    ws = wb.create_sheet(func_key)
    all_cols = ["generator", "type"] + FUNC_PARAM_COLS + ["x_min", "x_max", "source"]

    for col, h in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col, value=h)
        _style_header(cell)

    generators: dict = profile.get("generators", {})
    for r, (gen_name, gen_cfg) in enumerate(generators.items(), 2):
        func: dict = gen_cfg.get(func_key, {})
        params: dict = func.get("params", {})
        ws.cell(row=r, column=1, value=gen_name)
        ws.cell(row=r, column=2, value=func.get("type", ""))
        for col, pk in enumerate(FUNC_PARAM_COLS, 3):
            ws.cell(row=r, column=col, value=params.get(pk, ""))
        offset = 3 + len(FUNC_PARAM_COLS)
        ws.cell(row=r, column=offset, value=func.get("x_min", ""))
        ws.cell(row=r, column=offset + 1, value=func.get("x_max", ""))
        ws.cell(row=r, column=offset + 2, value=func.get("source", ""))

    _autofit(ws)


def _write_ess_sheet(wb: Workbook, profile: dict) -> None:
    ws = wb.create_sheet("ess")
    for col, h in enumerate(["parameter", "value"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        _style_header(cell)

    ess: dict = profile.get("ess", {})
    rows: list[tuple[str, Any]] = []

    if "short_dur" in ess:
        short = ess["short_dur"]
        for k, v in short.items():
            if not isinstance(v, dict):
                rows.append((f"short_dur.{k}", v))
    else:
        # legacy flat
        for k, v in ess.items():
            if k != "requirement_func" and not isinstance(v, dict):
                rows.append((k, v))
        req_func = ess.get("requirement_func", {})
        rows.append(("requirement_func_type", req_func.get("type", "")))
        for pk, pv in req_func.get("params", {}).items():
            rows.append((f"requirement_func_{pk}", pv))

    if "long_dur" in ess:
        long = ess["long_dur"]
        for k, v in long.items():
            if k == "requirement_func":
                rows.append(("long_dur.requirement_func_type", v.get("type", "")))
                for pk, pv in v.get("params", {}).items():
                    rows.append((f"long_dur.requirement_func_{pk}", pv))
            elif not isinstance(v, dict):
                rows.append((f"long_dur.{k}", v))

    for r, (k, v) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)

    _autofit(ws)


# ---------------------------------------------------------------------------
# Workbook → JSON profile
# ---------------------------------------------------------------------------

def workbook_to_profile(wb: Workbook) -> dict:
    profile: dict = {}
    profile.update(_read_country_sheet(wb))
    profile["generators"] = _read_generators_sheet(wb)
    for func_key in FUNC_SHEETS:
        func_data = _read_func_sheet(wb, func_key)
        for gen_name, func_cfg in func_data.items():
            if gen_name in profile["generators"]:
                profile["generators"][gen_name][func_key] = func_cfg
    profile["ess"] = _read_ess_sheet(wb)
    return profile


def _sheet_rows(wb: Workbook, sheet_name: str) -> list[list]:
    """Return all rows (including header) as lists of cell values."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def _read_country_sheet(wb: Workbook) -> dict:
    rows = _sheet_rows(wb, "country")
    if not rows:
        return {}
    result: dict = {}
    sources: list[str] = []
    for row in rows[1:]:  # skip header
        if not row or row[0] is None:
            continue
        key, value = str(row[0]), row[1]
        if key.startswith("source_"):
            if value:
                sources.append(str(value))
        elif key == "name":
            result["name"] = str(value) if value is not None else ""
        elif key == "annual_generation_twh":
            result["annual_generation_twh"] = float(value) if value is not None else 0
        elif key == "discount_rate":
            result["discount_rate"] = float(value) if value is not None else 0.05
    if sources:
        result["sources"] = sources
    return result


def _read_generators_sheet(wb: Workbook) -> dict:
    rows = _sheet_rows(wb, "generators")
    if not rows:
        return {}
    header = [str(c) if c is not None else "" for c in rows[0]]
    result: dict = {}
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        gen_name = str(row[0])
        cfg: dict = {}
        for col_idx, field in enumerate(header[1:], 1):
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None and val != "":
                cfg[field] = float(val) if isinstance(val, (int, float)) else val
        result[gen_name] = cfg
    return result


def _read_func_sheet(wb: Workbook, func_key: str) -> dict:
    rows = _sheet_rows(wb, func_key)
    if not rows:
        return {}
    header = [str(c) if c is not None else "" for c in rows[0]]
    result: dict = {}

    try:
        type_col = header.index("type")
        x_min_col = header.index("x_min")
        x_max_col = header.index("x_max")
        source_col = header.index("source")
    except ValueError:
        return result

    param_col_indices = {pk: i for i, pk in enumerate(header) if pk in FUNC_PARAM_COLS}

    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        gen_name = str(row[0])
        func_type = row[type_col] if type_col < len(row) else None
        if not func_type:
            continue

        params: dict = {}
        for pk, col_i in param_col_indices.items():
            val = row[col_i] if col_i < len(row) else None
            if val is not None and val != "":
                params[pk] = float(val)

        func_cfg: dict = {"type": func_type, "params": params}
        x_min = row[x_min_col] if x_min_col < len(row) else None
        x_max = row[x_max_col] if x_max_col < len(row) else None
        src = row[source_col] if source_col < len(row) else None
        if x_min is not None and x_min != "":
            func_cfg["x_min"] = float(x_min)
        if x_max is not None and x_max != "":
            func_cfg["x_max"] = float(x_max)
        if src:
            func_cfg["source"] = str(src)

        result[gen_name] = func_cfg

    return result


def _read_ess_sheet(wb: Workbook) -> dict:
    rows = _sheet_rows(wb, "ess")
    if not rows:
        return {}

    short: dict = {}
    long: dict = {}
    long_req_type = None
    long_req_params: dict = {}
    legacy: dict = {}
    legacy_req_type = None
    legacy_req_params: dict = {}

    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        key, value = str(row[0]), row[1]
        if value is None or value == "":
            continue
        val = float(value) if isinstance(value, (int, float)) else value

        if key.startswith("short_dur."):
            short[key[len("short_dur."):]] = val
        elif key.startswith("long_dur.requirement_func_type"):
            long_req_type = str(value)
        elif key.startswith("long_dur.requirement_func_"):
            pk = key[len("long_dur.requirement_func_"):]
            long_req_params[pk] = float(value)
        elif key.startswith("long_dur."):
            long[key[len("long_dur."):]] = val
        elif key == "requirement_func_type":
            legacy_req_type = str(value)
        elif key.startswith("requirement_func_"):
            pk = key[len("requirement_func_"):]
            legacy_req_params[pk] = float(value)
        else:
            legacy[key] = val

    if short or long:
        result: dict = {}
        if short:
            result["short_dur"] = short
        if long:
            if long_req_type:
                long["requirement_func"] = {"type": long_req_type, "params": long_req_params}
            result["long_dur"] = long
        return result
    else:
        if legacy_req_type:
            legacy["requirement_func"] = {"type": legacy_req_type, "params": legacy_req_params}
        return legacy


# ---------------------------------------------------------------------------
# File helpers
# ---------------------------------------------------------------------------

DATA_DIR = Path(__file__).parent.parent / "data" / "country_profiles"


def profile_path(code: str) -> Path:
    return DATA_DIR / f"{code.upper()}.json"


def excel_path(code: str) -> Path:
    return DATA_DIR / f"{code.upper()}.xlsx"


def load_profile_json(code: str) -> dict:
    path = profile_path(code)
    if not path.exists():
        raise FileNotFoundError(f"No profile found for country: {code}")
    return json.loads(path.read_text())


def save_profile_json(code: str, profile: dict) -> None:
    profile_path(code).write_text(json.dumps(profile, indent=2))


def generate_excel_from_json(code: str) -> None:
    """One-shot: read JSON profile → write Excel file next to it."""
    profile = load_profile_json(code)
    wb = profile_to_workbook(profile)
    wb.save(excel_path(code))


def load_excel_as_profile(code: str) -> dict:
    path = excel_path(code)
    if not path.exists():
        raise FileNotFoundError(f"No Excel profile for {code}. Generate it first.")
    wb = openpyxl.load_workbook(path)
    return workbook_to_profile(wb)


def save_uploaded_excel(code: str, file_bytes: bytes) -> dict:
    """Save uploaded Excel bytes, parse and return the resulting profile dict."""
    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    profile = workbook_to_profile(wb)
    # Persist both Excel and JSON
    excel_path(code).write_bytes(file_bytes)
    save_profile_json(code, profile)
    return profile
